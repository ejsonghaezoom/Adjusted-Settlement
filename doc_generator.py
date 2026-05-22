import os
import pandas as pd
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# 한글 폰트 설정 (Windows 기본 맑은 고딕 사용 시도, 없으면 기본 폰트)
try:
    pdfmetrics.registerFont(TTFont('MalgunGothic', 'malgun.ttf'))
    FONT_NAME = 'MalgunGothic'
except Exception:
    FONT_NAME = 'Helvetica' # 한글 깨짐 방지를 위해 시스템에 맞는 폰트 설치 필요

def format_currency(value):
    try:
        return f"{int(value):,}"
    except (ValueError, TypeError):
        return value

def generate_pdf(summary: dict, output_path: str):
    doc = SimpleDocTemplate(output_path, pagesize=A4,
                            rightMargin=40, leftMargin=40,
                            topMargin=40, bottomMargin=40)
    elements = []
    
    styles = getSampleStyleSheet()
    # 폰트 스타일 추가
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontName=FONT_NAME,
        fontSize=18,
        alignment=1, # Center
        spaceAfter=20
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontName=FONT_NAME,
        fontSize=10,
        spaceAfter=10
    )
    
    small_style = ParagraphStyle(
        'CustomSmall',
        parent=styles['Normal'],
        fontName=FONT_NAME,
        fontSize=8,
        spaceAfter=5
    )

    # 1. 헤더
    today = datetime.now()
    year_month = today.strftime('%Y년 %m월')
    
    plant_name = summary.get('발전소명', '발전소명')
    elements.append(Paragraph(f"(주){plant_name}", normal_style))
    elements.append(Paragraph(f"{year_month} 수정 정산 내역서", title_style))
    elements.append(Spacer(1, 12))

    # 2. 요약 정보
    summary_text = f"""
    <b>합계:</b> {format_currency(summary.get('총_차액', 0))} 원<br/>
    <b>공급가액:</b> {format_currency(summary.get('공급가액', 0))} 원<br/>
    <b>VAT:</b> {format_currency(summary.get('VAT', 0))} 원<br/>
    <b>작성일:</b> {today.strftime('%Y-%m-%d')}<br/>
    <b>발행 메일:</b> vpp.billing@haezoom.com<br/>
    <b>발행 가능 기간:</b> {summary.get('시작일', '-')} ~ {summary.get('종료일', '-')}<br/>
    <font size="8" color="gray">* 사업자번호 복수 소유 시 사업자번호기준 발행</font>
    """
    elements.append(Paragraph(summary_text, normal_style))
    elements.append(Spacer(1, 20))

    # 3. 테이블 1 (전력거래대금 세부 내역)
    data1 = [
        ['대상기간', '개소', '계량발전량(kWh)', '제어량(kWh)', '정산금(원)'],
        [
            f"{summary.get('시작일', '-')} ~ {summary.get('종료일', '-')}",
            f"{summary.get('데이터_개수', 0)}개",
            f"{format_currency(summary.get('총_계량발전량_차액', 0))}",
            f"{format_currency(summary.get('총_제어량_차액', 0))}",
            f"{format_currency(summary.get('총_차액', 0))}"
        ]
    ]
    t1 = Table(data1, colWidths=[130, 50, 100, 100, 100])
    t1.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), FONT_NAME),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(t1)
    elements.append(Spacer(1, 30))

    # 4. 테이블 2 (전력거래대금 세부항목)
    data2 = [
        ['항목', '금액(원)'],
        ['1. 전력량 정산금', format_currency(summary.get('전력량정산금_차액', 0))],
        ['2. 해줌 보상금', '0'], # 임시 0 처리
        ['3. 추가지급금', format_currency(summary.get('추가지급금_차액', 0))],
        ['4. 전력거래 수수료', '0'] # 임시 0 처리
    ]
    t2 = Table(data2, colWidths=[200, 150])
    t2.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -1), FONT_NAME),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(t2)
    elements.append(Spacer(1, 50))

    # 5. 푸터
    footer_text = """
    <b>HAEZOOM 로고</b><br/>
    주소: 서울특별시 송파구 법원로 128 SKV1 GL메트로시티 A동 706호<br/>
    주식회사 해줌 | 대표이사 권오현 | 연락처: 02-889-9941
    """
    elements.append(Paragraph(footer_text, small_style))

    doc.build(elements)
    return output_path


def generate_excel(df: pd.DataFrame, output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "수정정산결과"

    # 헤더 서식 지정
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_font = Font(bold=True)
    
    # '차이' 행 서식
    diff_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
    
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            
            if r_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                # '차이' 행 강조
                if df.iloc[r_idx-2]['구분'] == '차이':
                    cell.fill = diff_fill
                
                # 숫자 포맷 (천 단위 콤마)
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0'

    # 컬럼 너비 자동 조정
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(output_path)
    return output_path
