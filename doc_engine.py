import os
import openpyxl
from openpyxl.styles import PatternFill
from reportlab.pdfgen import canvas
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics

class DocumentGenerator:
    def __init__(self, output_dir="output_files"):
        self.output_dir = output_dir
        os.makedirs(self.output_dir, exist_ok=True)
        # Register font if needed (requires a ttf file like malgun.ttf)
        # pdfmetrics.registerFont(TTFont('Malgun', 'malgun.ttf'))

    def generate_pdf(self, plant_name, period_str, monthly_diff_data):
        file_name = f"{period_str} 수정정산 내역서_{plant_name}.pdf"
        file_path = os.path.join(self.output_dir, file_name)
        
        c = canvas.Canvas(file_path)
        # c.setFont('Malgun', 12)
        
        c.drawString(50, 800, f"[{plant_name}] {period_str} 수정 정산 내역서")
        c.drawString(50, 770, "작성일: 2026년 06월 02일")
        c.drawString(50, 750, "발행 메일: vpp.billing@haezoom.com")
        
        c.drawString(50, 700, "[ 전력거래대금 세부 내역 ]")
        c.drawString(70, 670, f"1. 전력량 정산금 차액: {monthly_diff_data.get('전력량정산금_차이', 0):,} 원")
        c.drawString(70, 650, f"2. 해줌 보상금 차액: {monthly_diff_data.get('해줌보상금_차이', 0):,} 원")
        c.drawString(70, 630, f"3. 추가지급금 차액: {monthly_diff_data.get('추가지급금_차이', 0):,} 원")
        c.drawString(70, 610, f"4. 전력거래 수수료 차액: {monthly_diff_data.get('전력거래수수료_차이', 0):,} 원")
        
        c.save()
        return file_path

    def generate_excel(self, plant_name, period_str, daily_diff_df):
        file_name = f"{period_str} 수정정산 상세내역_{plant_name}.xlsx"
        file_path = os.path.join(self.output_dir, file_name)
        
        plant_df = daily_diff_df[daily_diff_df['발전소명'] == plant_name]
        plant_df.to_excel(file_path, index=False)
        
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        highlight_fill = PatternFill(start_color="FFEAEA", end_color="FFEAEA", fill_type="solid")
        
        type_col_idx = None
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value == "구분":
                type_col_idx = col_idx
                break
                
        if type_col_idx:
            for row in ws.iter_rows(min_row=2):
                if row[type_col_idx - 1].value == "차이":
                    for cell in row:
                        cell.fill = highlight_fill
                        
        wb.save(file_path)
        return file_path
